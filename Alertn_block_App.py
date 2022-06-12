# -*- coding: utf-8 -*-
"""
Created on Fri Jun 10 20:37:08 2022

@author: ilias
"""
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd
import datetime
from io import BytesIO
import numpy as np

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(layout = "wide")

st.markdown("""Alternating Block method - Greece""")

col2, space2, col3 = st.columns((10,1,10))


stations = pd.read_excel('../Data/GR_StationsIDF_FD_English.xlsx')
stations_names = stations.loc[:,'name'].values.tolist()

with col2:
    
    #st.markdown("""Select rainfall station""", selected_station)
    selected_name = st.selectbox(label='Select category', options=stations_names)
    selected_station = stations.loc[stations['name'] == selected_name]
   
    # select return period
    min_T, mean_T, max_T = 5, 50, 100
    selected_T = st.slider("Select Return Period (T)", min_value=min_T,
    max_value=max_T, value=mean_T, step=5)
    
    # select time step
    min_dur, mean_dur, max_dur = 5, 10, 20
    sel_time_step = st.slider("Select time step (minutes)", min_value=min_dur,
    max_value=max_dur, value=mean_dur, step=5)
    
     # select storm duration
    storm_dur_options = np.arange(1,37,1)
    selected_storm_dur = st.selectbox(label='Select storm duration (h)',
                                options=storm_dur_options)
    
with col3:
    
    kappa = selected_station['kappa'].values[0]
    lambda_value = selected_station['lambda'].values[0]
    psi = selected_station['psi'].values[0]
    theta = selected_station['theta'].values[0]
    itta = selected_station['itta'].values[0]
    
    def rain_intens(T,d):
        a = lambda_value * (T**kappa - psi)
        b = (1 + d / theta) ** itta
        return a / b

    rain_intens_v = np.vectorize(rain_intens)
    d = np.linspace(0,24,100)
    
    st.markdown('Parameters of the station')
    st.write('kappa=',np.round(kappa,2),'\t', 'lambda=', np.round(lambda_value,2),
             'psi=',np.round(psi,2))
    st.write('theta=',np.round(theta,2),'itta=',np.round(itta,2))

    
    # IDF chart
    fig = go.Figure(layout=go.Layout(
        title=go.layout.Title(text="IDF Curves - "+str(selected_station['name'].values.tolist()[0]))))

    y = rain_intens_v(selected_T,d)
    fig.add_scatter(x=d, y=y, name='T='+str(selected_T))

    fig.update_layout(
        xaxis_title="Duration (h)",
        yaxis_title="Rainfall Intensity (mm/h)",
        legend_title="Return Periods")
    
    st.plotly_chart(fig, use_container_width=True)
    
    
    # Alternating block method
    # Pre-processing
    timestep_h = sel_time_step / 60
    # t is an array of the time steps needed in hours
    t = np.linspace(timestep_h, selected_storm_dur, int(selected_storm_dur/timestep_h))
    rain_i = rain_intens_v(selected_T, t) * t # rain intensity
    
    storm_duration_min = selected_storm_dur*60
    tot_timesteps_min = storm_duration_min / sel_time_step
    timesteps_min = np.linspace(sel_time_step, storm_duration_min, int(tot_timesteps_min))
    
    rain_intens_1st_step = np.array([rain_i[0]])
    # I also kept the first value besides teh differences
    differences = np.diff(rain_i)
    rain_intens_steps = np.concatenate((rain_intens_1st_step,differences))
    merge_array = np.column_stack((rain_intens_steps, timesteps_min))
    rain_steps_df = pd.DataFrame(data = merge_array, index=t,
                             columns=['precip', 'steps'])
    
    
    # Processing - Alternative block method
    atltern_array = np.zeros(rain_steps_df.shape[0]) 
    rain_steps_df_sorted = rain_steps_df.sort_values(by='precip', ascending=False)
    mean_index = int(rain_steps_df.shape[0] /2)
    
    switch = True
    l1 = 1 
    l2 = 1
    for i in range(rain_steps_df_sorted.shape[0]):
        if  i == 0:
            atltern_array[mean_index] = rain_steps_df_sorted['precip'].iloc[0]
        else:
            if switch == True:
                atltern_array[mean_index - l1] = rain_steps_df_sorted['precip'].iloc[i]
                switch = False
                l1+=1
            else:
                atltern_array[mean_index + l2] = rain_steps_df_sorted['precip'].iloc[i]
                switch = True
                l2+=1            
    
    #atltern_df = pd.DataFrame(atltern_array, index=np.arange(1,rain_steps_df_sorted.shape[0]+1,1), columns= ['precip'])
    atltern_ar_merge = np.column_stack((atltern_array, timesteps_min))
    atltern_df = pd.DataFrame(atltern_ar_merge, index=t, columns= ['precip', 'time_ste_h'])
    atltern_excel = atltern_df.to_excel('Alternating_block_method_app.xlsx', float_format='%.2f')
    
    # Export to excel
    
    
    # Show plot
    fig_Altern = go.Figure(layout=go.Layout(
        title=go.layout.Title(text="Alternating block method ")))
    fig_Altern.add_scatter(x=atltern_df.index, y=atltern_df.precip, name='T='+str(selected_T))

    fig_Altern.update_layout(
        xaxis_title="Duration (h)",
        yaxis_title="Rainfall Dh (mm)")
    
    st.plotly_chart(fig_Altern, use_container_width=True)
    
    
    # Excel output configuration
    wb = load_workbook('Alternating_block_method_app.xlsx')
    ws = wb['Sheet1']
    ws.title = 'Alternating_block'
    ws = wb['Alternating_block']
    wb.create_sheet('Plot')
    ws['D3'] = 'Station'
    ws['D3'].font = Font(name="Arial", size=14, color="00FF0000")
    ws['D3'].alignment = Alignment(horizontal='center')
    ws['D4'] = str(selected_station['name'].values[0])
    
    # applying border style
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    max_row = wb['Alternating_block'].max_row
    for j in range(1,3):
        for i in range(1,max_row):
            ws.cell(column=j, row=i).border = thin_border
    
    ws.merge_cells('D4:E4')
    
    # Descriptive statitsics
    ws['G2'] = 'Descriptive Statistics'
    ws['G2'].alignment = Alignment(horizontal='center')
    ws['G2'].font = Font(bold=True)
    names = ['min', 'max', 'mean', 'std']
    
    column = 7
    for i, value in enumerate(names):
        cell = ws.cell(column=column, row=3+i, value=value)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='16A085', end_color='16A085', fill_type="solid")
        cell.font = Font(bold=True)
        
    # Descriptive stats
    ws['H3'].value = round(atltern_df.precip.min(),2)
    ws['H4'].value = round(atltern_df.precip.max(),2)
    ws['H5'].value = round(atltern_df.precip.mean(),2)
    ws['H6'].value = round(np.std(atltern_df.precip),2)
    
    # Number format
    rows = 26
    for row in range(1, rows):
        ws["B{}".format(row)].number_format = '#,##0.00'
    
    # Adjusting column width size
    column_widths = []
    for row in wb['Alternating_block'].iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(str(cell.value)))
    
    for i, column_width in enumerate(column_widths):
        wb['Alternating_block'].column_dimensions[get_column_letter(i + 1)].width = column_width    
    
    

    
# Creating a barchart
chart1 = BarChart()
data = Reference(ws, min_col = 2, min_row = 1, max_row = max_row) # You need to include the name of the column as well
# besides the data
cats = Reference(ws, min_col = 1, min_row = 2, max_row = max_row)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
chart1.title = "Alternating_Block_Method - Hyetograph"
chart1.x_axis.title = 'Time (h)'
chart1.y_axis.title = 'Rainfall (mm)'
wb['Plot'].add_chart(chart1, "C4")


wb['Alternating_block'].insert_rows(1)
# Freeze header
wb["Alternating_block"].freeze_panes = "C2"
# define printing area
wb['Alternating_block'].print_area = "A1:I27"

wb.save('Alternating_block_method_app.xlsx')
  
# Download excel file     
with open('Alternating_block_method_app.xlsx', 'rb') as f:
    st.download_button('Download Excel file', f, file_name = 'Altern_block.xlsx',
                       mime = 'application/vnd.ms-excel')    
    
# Download image file


# Download pdf file


# add map of the station  - SOS
# fig_map = px.scatter_mapbox(df_clip_2, lon='x', lat='y',
#                     color_discrete_sequence=["fuchsia"], hover_data=["Date", "Time"], zoom=9, height=500)
# fig_map.update_layout(mapbox_style="open-street-map")
        
    
    
    